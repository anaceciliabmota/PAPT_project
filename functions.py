#create read file function

#read preferences using


from openpyxl import load_workbook
from structs import *
from collections import namedtuple

def readClasses(file, professor_dic):

    class_sheet = file["InfoTurmas"]

    last_line = len(class_sheet["A"])

    #read class features
    list_of_classes = []

    for line in range(2, last_line+1):
        code = class_sheet.cell(row=line, column=1).value
        name = class_sheet.cell(row=line, column=2).value
        number = class_sheet.cell(row=line, column=3).value
        credit = class_sheet.cell(row=line, column=4).value
        time = class_sheet.cell(row=line, column=5).value
        offered = class_sheet.cell(row=line, column=6).value
        center = class_sheet.cell(row=line, column=7).value
        pre_alocation = class_sheet.cell(row=line, column=8).value


        if not offered:
            break
        
        if offered == "SIM":
            offered = 1
        else:
            offered = 0

        if not code and offered:
            break
        
        if not name and offered:
            break

        if not number and offered:
            break

        if not credit and offered:
            break
        
        if not center and offered:
            break

        if not time and offered:
            break
        
        if pre_alocation:
            pre_alocation = professor_dic[pre_alocation].index
        else:
            pre_alocation = 0
        
        if time: 
            time = time_converter(time)

        if offered:
            t = Class(name, number, code, credit, time, center, pre_alocation, len(list_of_classes))
        
            list_of_classes.append(t)

    return list_of_classes

def readFile(filename):

    try:
        
        file = load_workbook(filename)
        
    except FileNotFoundError:
        
        print(f"Erro: O arquivo '{filename}' não foi encontrado.")
        exit(0)

    except Exception as e:
            
        print(f"Ocorreu um erro ao tentar carregar o arquivo: {e}")
        exit(0)
        
    list_of_professors = readProfessor(file)

    professor_dic = {}
    classes_dic = {}

    for p in list_of_professors:
        professor_dic[p.name] = p

    list_of_classes = readClasses(file, professor_dic)

    readPreferencesClasses(professor_dic, list_of_classes, file)

    penalty_map = readPenalty(file)

    preferences_map = readPreferenceValues(file)

    return list_of_classes, list_of_professors, preferences_map, penalty_map


def readPenalty(file):
    penalty_sheet = file["Penalidades"]

    penalty = {}

    days = penalty_sheet.cell(row=2, column=2).value
    penalty["dias"] = days
    shift = penalty_sheet.cell(row=3, column=2).value
    penalty["turno"] = shift
    pending_class =  penalty_sheet.cell(row=4, column=2).value
    penalty["pendente"] = pending_class
    diff = penalty_sheet.cell(row=5, column=2).value
    penalty["diff"] = diff
    overload = penalty_sheet.cell(row=6, column=2).value
    penalty["sobrecarga"] = overload
    min_credits = penalty_sheet.cell(row=7, column=2).value
    penalty["min"] = min_credits
    max_credits = penalty_sheet.cell(row=8, column=2).value
    penalty["max"] = max_credits
    work_at_night = penalty_sheet.cell(row=9, column=2).value
    penalty["noturno"] = work_at_night

    return penalty



def readPreferenceValues(file):
    preference_sheet = file["Preferencias_Beneficios"]

    last_line = len(preference_sheet["A"])

    preferences = {}

    for line in range(2, last_line+1):
        #p is the preference number
        p = preference_sheet.cell(row=line, column=1).value
        # value is the associated value of the number in the model
        value = preference_sheet.cell(row=line, column=2).value
        
        preferences[p] = value
    
    return preferences

def readProfessor(file):
    professor_sheet = file["Professores"]

    last_line = len(professor_sheet["A"])

    professor_list = []

    for line in range(2, last_line+1):
        name = professor_sheet.cell(row=line, column=1).value
        blocked_time = professor_sheet.cell(row=line, column=2).value
        min_credits = professor_sheet.cell(row=line, column=3).value
        max_credits = professor_sheet.cell(row=line, column=4).value
        daily_max_credits = professor_sheet.cell(row=line, column=5).value
        max_class_days = professor_sheet.cell(row=line, column=6).value

        if not name:
            break
        
        if not max_credits:
            break

        if not daily_max_credits:
            break

        if not max_class_days:
            break

        #must change the time to a number and verify if time exist

        p = Professor(name, blocked_time, min_credits, max_credits, daily_max_credits, max_class_days, len(professor_list))

        professor_list.append(p)

    return professor_list

def readPreferencesClasses(professor_dic, list_of_classes, file):
    class_sheet = file["InfoTurmas"]
    
    # which is better: class_index or class object
    Preferences = namedtuple('Preferences', ['class_index', 'value'])

    last_line = len(class_sheet["A"])

    class_counter = 0

    for i in range(9, 30+9):
        
        preference_list = []
        professor = class_sheet.cell(row=1, column=i).value
        if not professor:
            break
        professor = professor_dic[professor]

        for line in range(2, last_line):
            offer = class_sheet.cell(row=line, column=6).value
            
            if offer == "NÃO":
                continue
            else:
                pref_value = class_sheet.cell(row=line, column=i).value

                if pref_value: 
                    pair = Preferences(class_counter, pref_value)
                    preference_list.append(pair)
            
            class_counter += 1
            
        class_counter = 0

        professor.set_preference(preference_list)

def time_converter(time):

    time_list = time.split()

    times = []
    # binary list representing the times: 1 if the class has this time, 0 if it doesn't.
    
    binary_list = [0 for _ in range (96)]

    for time in time_list:

        first_part = True
        days = []
        hour = []

        for t in time:
            
            if first_part and t.isdigit():
                days.append(t)
            elif t.isalpha():
                turno = t
                first_part = False
            else:
                hour.append(t)


        for d in days:
            for h in hour:
                times.append(d + turno + h)
        
    # chance the time representation to a number from 1 to 96

    for i in range(len(times)):
        day = int(times[i][0])
        shift = times[i][1]
        hour = int(times[i][2])
        if shift == "T":
            index = (day - 2) * 16 + 6 + hour
        elif shift == "N":
            index = (day - 2) * 16 + 12 + hour
        else:
            index = (day - 2) * 16 + hour
        
        binary_list[index-1] = 1
    
        
    return binary_list





